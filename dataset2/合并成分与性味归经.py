# import csv
#
# # 读取txt文件中的中药成分数据
# herb_components = {}
# with open('herb_browse_final_columns_removed.txt', 'r', encoding='utf-8') as txt_file:
#     reader = csv.reader(txt_file, delimiter='\t')
#     next(reader)  # 跳过表头
#     for row in reader:
#         if len(row) >= 2:
#             herb_name = row[0].strip()
#             components = row[1].strip()
#             herb_components[herb_name] = components
#
# # 读取并更新csv文件
# rows = []
# with open('气味归经去重.csv', 'r', encoding='utf-8') as csv_file:
#     reader = csv.reader(csv_file)
#     header = next(reader)  # 读取表头
#     rows.append(header)
#
#     for row in reader:
#         herb_name = row[0].strip()
#         # 获取成分，如果没有找到则留空
#         components = herb_components.get(herb_name, '')
#         # 确保行有足够的列
#         while len(row) < 5:
#             row.append('')
#         # 更新第五列（索引为4）
#         if len(row) > 4:
#             row[4] = components
#         else:
#             row.append(components)
#         rows.append(row)
#
# # 写回csv文件
# with open('气味归经去重.csv', 'w', encoding='utf-8', newline='') as csv_file:
#     writer = csv.writer(csv_file)
#     writer.writerows(rows)
#
# print("成分数据已成功写入csv文件的第五列")

import csv
from openpyxl import load_workbook, Workbook

# 读取txt文件中的中药成分数据
herb_components = {}
with open('herb_browse_final_columns_removed.txt', 'r', encoding='utf-8') as txt_file:
    reader = csv.reader(txt_file, delimiter='\t')
    next(reader)  # 跳过表头
    for row in reader:
        if len(row) >= 2:
            herb_name = row[0].strip()
            components = row[1].strip()
            herb_components[herb_name] = components

# 读取Excel文件
wb = load_workbook('气味归经去重.xlsx')
ws = wb.active

# 更新Excel文件
for row in ws.iter_rows(min_row=2):  # 从第2行开始（跳过表头）
    herb_name = row[0].value.strip() if row[0].value else ""
    components = herb_components.get(herb_name, '')

    # 确保有足够的列（至少5列）
    if len(row) < 5:
        for _ in range(5 - len(row)):
            ws.cell(row=row[0].row, column=len(row) + 1, value="")

    # 更新第5列（E列）
    ws.cell(row=row[0].row, column=5, value=components)

# 保存修改后的Excel文件
wb.save('气味归经去重_更新.xlsx')
print("成分数据已成功写入Excel文件的第五列")
